# error_history/core/manager.py
"""
错误历史持久化子系统 - 核心管理器
"""

import sqlite3
import json
import logging
import threading
import time
import os
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any, Tuple
from contextlib import contextmanager
import builtins

from .models import ErrorRecord, DailyStatistics, ErrorHistoryConfig, ErrorSeverity, ErrorCategory

logger = logging.getLogger(__name__)


class ErrorHistoryManager:
    """错误历史管理器 - 负责数据库操作和数据管理"""

    def __init__(self, db_path: str = None, config_manager = None):
        """
        初始化错误历史管理器

        Args:
            db_path: 数据库文件路径
            config_manager: 配置管理器实例
        """
        self.config_manager = config_manager
        self.db_path = Path(db_path) if db_path else self._get_default_db_path()
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

        # 数据库连接池
        self._connection_pool = {}
        self._pool_lock = threading.Lock()
        self._max_connections = 20  # 增加连接池大小

        # 调度/监听线程控制
        self._cleanup_thread = None
        self._cleanup_stop_event = threading.Event()
        self._last_cleanup_run_ts = 0.0
        self._config_watch_thread = None
        self._config_watch_stop_event = threading.Event()
        self._watched_files = {}

        # 配置
        self.config = ErrorHistoryConfig()

        # 初始化
        self._init_database()
        self._load_config()
        self._setup_config_listeners()

        # 启动配置文件热重载监听与自动清理调度
        try:
            self._start_config_watcher()
        except Exception as e:
            logger.debug(f"启动配置热重载监听失败: {e}")
        # 可选启用 watchdog 事件监听（存在则使用）
        try:
            self._try_watchdog_start()
        except Exception as e:
            logger.debug(f"watchdog监听不可用: {e}")
        try:
            self._start_cleanup_scheduler()
        except Exception as e:
            logger.debug(f"启动自动清理调度失败: {e}")

        logger.info(f"错误历史管理器初始化完成，数据库: {self.db_path}")

    def _get_default_db_path(self) -> Path:
        """获取默认数据库路径"""
        return Path("data/error_history.db")

    def _load_config(self):
        """加载配置"""
        try:
            if self.config_manager:
                loaded = False
                # 1) 优先：features/error_history.json（与 README 一致，嵌套结构）
                try:
                    data_fea = self.config_manager._load_config_file('features/error_history')
                    if data_fea:
                        self.config = ErrorHistoryConfig.from_dict(data_fea)
                        logger.info("从 features/error_history.json 加载错误历史配置")
                        loaded = True
                except Exception as exc:
                    logger.debug(f"读取 features/error_history.json 失败: {exc}")

                # 2) 次选：error_handling.json（011任务扁平结构，需映射）
                if not loaded:
                    try:
                        data_eh = self.config_manager._load_config_file('error_handling')
                        if data_eh and isinstance(data_eh, dict) and 'error_history' in data_eh:
                            raw = data_eh.get('error_history', {}) or {}
                            nested = {
                                'version': raw.get('version', '1.0'),
                                'enabled': raw.get('enabled', True),
                                'database': {
                                    'path': raw.get('database_path', 'data/error_history.db'),
                                    'max_connections': raw.get('max_connections', 5),
                                    'timeout_seconds': raw.get('timeout_seconds', 30),
                                    'backup_enabled': raw.get('backup_enabled', True),
                                    'backup_interval_hours': raw.get('backup_interval_hours', 24),
                                },
                                'retention': {
                                    'days': raw.get('retention_days', 90),
                                    'auto_cleanup': raw.get('auto_cleanup', True),
                                    'cleanup_schedule': raw.get('cleanup_schedule', '0 2 * * *'),
                                    'compression_enabled': raw.get('compression_enabled', False)
                                },
                                'ui': {},
                                'monitoring': {},
                                'export': {}
                            }
                            self.config = ErrorHistoryConfig.from_dict(nested)
                            logger.info("从 error_handling.json(扁平) 加载错误历史配置")
                            loaded = True
                    except Exception as exc:
                        logger.debug(f"读取 error_handling.json 失败: {exc}")

                # 3) 兼容旧实现：get_config('features') 路径
                if not loaded:
                    try:
                        config_data = self.config_manager.get_config("error_history", {}, "features")
                        if config_data:
                            self.config = ErrorHistoryConfig.from_dict(config_data)
                            logger.info("从配置管理器(features.error_history)加载错误历史配置")
                            loaded = True
                    except Exception as exc:
                        logger.debug(f"从配置管理器读取 features.error_history 失败: {exc}")

                if not loaded:
                    logger.warning("未找到错误历史配置，尝试从数据库加载配置")
                    self._load_config_from_db()
            else:
                logger.warning("配置管理器不可用，尝试从数据库加载配置")
                self._load_config_from_db()
        except Exception as e:
            logger.error(f"加载配置失败: {e}，使用默认配置")
        # 应用运行期配置（如连接池大小）
        try:
            self._apply_config_runtime()
        except Exception as e:
            logger.debug(f"应用运行期配置失败: {e}")

    def _load_config_from_db(self):
        """从本地数据库 system_config 表加载配置（若存在）。"""
        try:
            with self._get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    "SELECT config_value FROM system_config WHERE config_key = ?",
                    ("error_history",)
                )
                row = cursor.fetchone()
                if row and row[0]:
                    try:
                        data = json.loads(row[0])
                        self.config = ErrorHistoryConfig.from_dict(data)
                        logger.info("已从数据库加载错误历史配置")
                    except Exception as exc:
                        logger.warning(f"解析数据库配置失败，将使用默认配置: {exc}")
        except Exception as e:
            logger.debug(f"从数据库加载配置失败: {e}")

    def save_config(self) -> bool:
        """将当前配置持久化到数据库，并应用运行期参数。"""
        try:
            config_json = json.dumps(self.config.to_dict(), ensure_ascii=False)
            with self._get_connection() as conn:
                try:
                    conn.execute(
                        """
                        INSERT INTO system_config (config_key, config_value, config_type, description)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(config_key) DO UPDATE SET
                            config_value=excluded.config_value,
                            config_type=excluded.config_type,
                            description=excluded.description,
                            updated_at=CURRENT_TIMESTAMP
                        """,
                        ("error_history", config_json, "json", "Error History Config")
                    )
                except sqlite3.Error as upsert_err:
                    # 兼容旧版 SQLite：不支持 ON CONFLICT DO UPDATE，回退为 REPLACE INTO
                    try:
                        conn.execute(
                            "REPLACE INTO system_config (config_key, config_value, config_type, description) VALUES (?, ?, ?, ?)",
                            ("error_history", config_json, "json", "Error History Config")
                        )
                        logger.info(f"ON CONFLICT 不可用，已使用 REPLACE 回退：{upsert_err}")
                    except sqlite3.Error as replace_err:
                        raise replace_err
                conn.commit()
            # 应用运行期配置
            self._apply_config_runtime()
            logger.info("错误历史配置已保存并应用")
            return True
        except Exception as e:
            logger.error(f"保存错误历史配置失败: {e}")
            return False

    def _apply_config_runtime(self):
        """根据配置应用运行期参数（例如连接池最大连接数）。"""
        try:
            if isinstance(self.config.max_connections, int) and self.config.max_connections > 0:
                self._max_connections = int(self.config.max_connections)
        except Exception as e:
            logger.debug(f"应用运行期配置 _max_connections 失败: {e}")

        # 应用运行期：根据配置重启自动清理调度
        try:
            self._restart_cleanup_scheduler()
        except Exception as e:
            logger.debug(f"重启自动清理调度失败: {e}")

    def _setup_config_listeners(self):
        """设置配置监听器"""
        if self.config_manager:
            try:
                # 统一注册一次回调（该管理器的通知仅在保存 app/ui/file_types 时触发）
                self.config_manager.add_change_listener(self._on_config_changed)
                logger.info("已设置错误历史配置监听器")
            except Exception as e:
                logger.warning(f"设置配置监听器失败: {e}")

    def _on_config_changed(self, config_name: str):
        """配置变更处理（来自 ConfigManager 的简单通知）。"""
        # 由于 error_history 配置不在 app/ui/file_types 中，此处仅作为桥接，实际热重载由文件监控完成。
        logger.debug(f"收到配置变更通知: {config_name}")
        # 仍执行一次应用层面的运行期参数刷新（以便外部依赖变更时同步）。
        try:
            self._apply_config_runtime()
        except Exception:
            pass

    # ==================== 配置热重载（文件监控） ====================
    def _get_watch_files(self) -> List[Path]:
        files = []
        try:
            if self.config_manager and hasattr(self.config_manager, 'config_dir'):
                base = self.config_manager.config_dir
                files.append(base / 'features' / 'error_history.json')
                files.append(base / 'error_handling.json')
        except Exception:
            pass
        return [p for p in files if p is not None]

    def _start_config_watcher(self):
        if self._config_watch_thread and self._config_watch_thread.is_alive():
            return
        self._config_watch_stop_event.clear()
        # 初始化被监控文件及其mtime
        for p in self._get_watch_files():
            try:
                self._watched_files[str(p)] = os.path.getmtime(p) if p.exists() else 0.0
            except Exception:
                self._watched_files[str(p)] = 0.0

        def _loop():
            while not self._config_watch_stop_event.is_set():
                changed = False
                for path, last_mtime in list(self._watched_files.items()):
                    try:
                        mtime = os.path.getmtime(path) if os.path.exists(path) else 0.0
                    except Exception:
                        mtime = 0.0
                    if mtime != last_mtime:
                        self._watched_files[path] = mtime
                        changed = True
                if changed:
                    try:
                        logger.info("检测到错误历史配置文件变更，重新加载配置")
                        # 清理配置缓存，确保读取最新文件
                        try:
                            if self.config_manager and hasattr(self.config_manager, 'reload_config'):
                                self.config_manager.reload_config()
                        except Exception:
                            pass
                        self._load_config()
                    except Exception as e:
                        logger.warning(f"配置热重载失败: {e}")
                # 轮询间隔
                self._config_watch_stop_event.wait(2.0)

        self._config_watch_thread = threading.Thread(target=_loop, name="EHConfigWatcher", daemon=True)
        self._config_watch_thread.start()

    def _stop_config_watcher(self):
        try:
            if self._config_watch_thread and self._config_watch_thread.is_alive():
                self._config_watch_stop_event.set()
                self._config_watch_thread.join(timeout=3.0)
        except Exception:
            pass

    # 可选：watchdog 事件监听增强（存在即用）
    def _try_watchdog_start(self):
        try:
            import watchdog.events  # type: ignore
            import watchdog.observers  # type: ignore
        except Exception:
            return False
        try:
            base = getattr(self.config_manager, 'config_dir', None)
            if not base:
                return False
            base = Path(base)
            patterns = {str(base / 'features' / 'error_history.json'), str(base / 'error_handling.json')}

            class _Handler(watchdog.events.FileSystemEventHandler):  # type: ignore
                def __init__(self, outer):
                    self.outer = outer
                def on_modified(self, event):  # noqa
                    try:
                        if event.src_path.replace('\\','/') in [p.replace('\\','/') for p in patterns]:
                            if self.outer.config_manager and hasattr(self.outer.config_manager, 'reload_config'):
                                self.outer.config_manager.reload_config()
                            self.outer._load_config()
                    except Exception as e:
                        logger.debug(f"watchdog处理失败: {e}")

            self._watchdog_observer = watchdog.observers.Observer()  # type: ignore
            self._watchdog_handler = _Handler(self)
            self._watchdog_observer.schedule(self._watchdog_handler, str(base), recursive=True)
            self._watchdog_observer.daemon = True
            self._watchdog_observer.start()
            logger.info("已启用watchdog文件变更监听")
            return True
        except Exception as e:
            logger.debug(f"watchdog启动失败: {e}")
            return False

    def _try_watchdog_stop(self):
        try:
            obs = getattr(self, '_watchdog_observer', None)
            if obs:
                obs.stop()
                obs.join(timeout=3.0)
        except Exception:
            pass

    # ==================== 自动清理调度 ====================
    def _start_cleanup_scheduler(self):
        if self._cleanup_thread and self._cleanup_thread.is_alive():
            return
        if not getattr(self.config, 'auto_cleanup', False):
            return
        self._cleanup_stop_event.clear()

        def _loop():
            logger.info("自动清理调度器已启动")
            while not self._cleanup_stop_event.is_set():
                try:
                    interval = self._get_cleanup_interval_seconds()
                    if interval is not None:
                        # 间隔模式（如 @every 60s）
                        now = time.time()
                        if now - self._last_cleanup_run_ts >= max(1.0, interval):
                            self._run_cleanup_task()
                            self._last_cleanup_run_ts = time.time()
                        # 小步睡眠，便于快速响应停止
                        self._cleanup_stop_event.wait(0.5)
                        continue

                    # Cron 简化：m h * * *
                    h, m = self._parse_cleanup_cron()
                    now_dt = datetime.now()
                    # 每分钟检查一次
                    if now_dt.minute == m and now_dt.hour == h:
                        # 保证每日只执行一次
                        run_key = now_dt.strftime('%Y-%m-%d')
                        last_key = time.strftime('%Y-%m-%d', time.localtime(self._last_cleanup_run_ts)) if self._last_cleanup_run_ts else ''
                        if run_key != last_key:
                            self._run_cleanup_task()
                            self._last_cleanup_run_ts = time.time()
                    self._cleanup_stop_event.wait(10.0)
                except Exception as e:
                    logger.debug(f"自动清理调度循环异常: {e}")
                    self._cleanup_stop_event.wait(5.0)

        self._cleanup_thread = threading.Thread(target=_loop, name="EHCleanup", daemon=True)
        self._cleanup_thread.start()

    def _stop_cleanup_scheduler(self):
        try:
            if self._cleanup_thread and self._cleanup_thread.is_alive():
                self._cleanup_stop_event.set()
                self._cleanup_thread.join(timeout=3.0)
        except Exception:
            pass

    def _restart_cleanup_scheduler(self):
        self._stop_cleanup_scheduler()
        self._start_cleanup_scheduler()

    def _get_cleanup_interval_seconds(self) -> Optional[float]:
        """解析 '@every 60s' '@every 5m' '@every 1h' 风格的间隔，返回秒数或None。"""
        try:
            sched = getattr(self.config, 'cleanup_schedule', '') or ''
            s = str(sched).strip().lower()
            if not s.startswith('@every'):
                return None
            parts = s.split()
            if len(parts) != 2:
                return None
            val = parts[1]
            if val.endswith('ms') and val[:-2].isdigit():
                return max(0.001, float(val[:-2]) / 1000.0)
            if val.endswith('s') and val[:-1].isdigit():
                return float(val[:-1])
            if val.endswith('m') and val[:-1].isdigit():
                return float(val[:-1]) * 60.0
            if val.endswith('h') and val[:-1].isdigit():
                return float(val[:-1]) * 3600.0
            if val.isdigit():
                return float(val)
        except Exception:
            return None
        return None

    def _parse_cleanup_cron(self) -> Tuple[int, int]:
        """解析 'm h * * *' 格式，返回 (hour, minute)。失败则默认 (2,0)。"""
        try:
            sched = getattr(self.config, 'cleanup_schedule', '') or ''
            parts = str(sched).strip().split()
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                return int(parts[1]) % 24, int(parts[0]) % 60
        except Exception:
            pass
        return 2, 0

    def _run_cleanup_task(self):
        try:
            days = getattr(self.config, 'retention_days', 90) or 90
            deleted = self.cleanup_old_errors(days=days)
            logger.info(f"自动清理执行完成，清理 {deleted} 条记录（> {days} 天）")
            return deleted
        except Exception as e:
            logger.warning(f"自动清理任务执行失败: {e}")
            return 0

    # ==================== 公共接口 ====================
    def trigger_cleanup_now(self) -> int:
        try:
            return int(self._run_cleanup_task() or 0)
        except Exception:
            return 0

    def restart_scheduler(self) -> bool:
        try:
            self._restart_cleanup_scheduler()
            return True
        except Exception:
            return False

    def _compute_next_run_datetime(self) -> Optional[datetime]:
        try:
            interval = self._get_cleanup_interval_seconds()
            now = datetime.now()
            if interval is not None:
                base = datetime.fromtimestamp(self._last_cleanup_run_ts) if self._last_cleanup_run_ts else now
                return base + timedelta(seconds=max(1.0, interval))
            h, m = self._parse_cleanup_cron()
            candidate = now.replace(hour=h, minute=m, second=0, microsecond=0)
            if candidate <= now:
                candidate = candidate + timedelta(days=1)
            return candidate
        except Exception:
            return None

    def get_cleanup_status(self) -> Dict[str, Any]:
        try:
            next_dt = self._compute_next_run_datetime()
            return {
                'enabled': bool(getattr(self.config, 'auto_cleanup', False)),
                'running': bool(self._cleanup_thread and self._cleanup_thread.is_alive()),
                'last_run_ts': self._last_cleanup_run_ts,
                'last_run': datetime.fromtimestamp(self._last_cleanup_run_ts).isoformat() if self._last_cleanup_run_ts else None,
                'next_run': next_dt.isoformat() if next_dt else None,
                'schedule': getattr(self.config, 'cleanup_schedule', ''),
                'mode': 'interval' if self._get_cleanup_interval_seconds() is not None else 'daily'
            }
        except Exception:
            return {
                'enabled': False,
                'running': False,
                'last_run_ts': 0.0,
                'last_run': None,
                'next_run': None,
                'schedule': '',
                'mode': 'unknown'
            }

    @contextmanager
    def _get_connection(self):
        """获取数据库连接（连接池管理）"""
        thread_id = threading.get_ident()
        conn = None

        with self._pool_lock:
            # 尝试获取现有连接
            if thread_id in self._connection_pool:
                conn = self._connection_pool[thread_id]
                try:
                    # 测试连接是否有效
                    conn.execute("SELECT 1").fetchone()
                except sqlite3.Error:
                    # 连接无效，关闭并移除
                    try:
                        conn.close()
                    except:
                        pass
                    del self._connection_pool[thread_id]
                    conn = None

            # 创建新连接
            if conn is None and len(self._connection_pool) < self._max_connections:
                conn = sqlite3.connect(
                    str(self.db_path),
                    timeout=self.config.timeout_seconds,
                    check_same_thread=False
                )
                conn.execute("PRAGMA foreign_keys = ON")
                conn.execute("PRAGMA journal_mode = WAL")
                conn.execute("PRAGMA synchronous = NORMAL")
                conn.execute("PRAGMA cache_size = -64000")  # 64MB缓存
                conn.execute("PRAGMA temp_store = MEMORY")
                conn.execute("PRAGMA busy_timeout = 30000")  # 30秒忙等待超时

                self._connection_pool[thread_id] = conn

        if conn is None:
            # 如果连接池满了，尝试清理无效连接
            with self._pool_lock:
                # 清理无效连接
                invalid_connections = []
                for tid, pool_conn in self._connection_pool.items():
                    try:
                        pool_conn.execute("SELECT 1").fetchone()
                    except sqlite3.Error:
                        invalid_connections.append(tid)

                for tid in invalid_connections:
                    try:
                        self._connection_pool[tid].close()
                    except:
                        pass
                    del self._connection_pool[tid]

                # 再次尝试创建连接
                if len(self._connection_pool) < self._max_connections:
                    conn = sqlite3.connect(
                        str(self.db_path),
                        timeout=self.config.timeout_seconds,
                        check_same_thread=False
                    )
                    conn.execute("PRAGMA foreign_keys = ON")
                    conn.execute("PRAGMA journal_mode = WAL")
                    conn.execute("PRAGMA synchronous = NORMAL")
                    conn.execute("PRAGMA cache_size = -64000")
                    conn.execute("PRAGMA temp_store = MEMORY")
                    conn.execute("PRAGMA busy_timeout = 30000")

                    self._connection_pool[thread_id] = conn

        if conn is None:
            raise sqlite3.Error("无法获取数据库连接：连接池已满且无无效连接可清理")

        try:
            yield conn
        finally:
            # 连接由连接池管理，不在这里关闭
            pass

    def _init_database(self):
        """初始化数据库"""
        try:
            with self._get_connection() as conn:
                self._create_tables(conn)
                self._create_indexes(conn)
                logger.info("数据库初始化完成")
        except Exception as e:
            logger.error(f"数据库初始化失败: {e}")
            raise

    def _create_tables(self, conn: sqlite3.Connection):
        """创建数据表"""
        # 错误历史表
        conn.execute('''
            CREATE TABLE IF NOT EXISTS error_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                error_id TEXT NOT NULL UNIQUE,
                error_type TEXT NOT NULL,
                error_message TEXT NOT NULL,
                severity TEXT NOT NULL,
                category TEXT NOT NULL,
                module TEXT,
                function TEXT,
                line_number INTEGER,
                stack_trace TEXT,
                context TEXT,
                user_context TEXT,
                system_context TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                resolved BOOLEAN DEFAULT 0,
                resolved_at TIMESTAMP,
                resolution_method TEXT,
                resolution_time REAL,
                retry_count INTEGER DEFAULT 0,
                max_retries INTEGER DEFAULT 3,
                tags TEXT,
                metadata TEXT
            )
        ''')

        # 日统计表
        conn.execute('''
            CREATE TABLE IF NOT EXISTS error_statistics_daily (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date DATE NOT NULL UNIQUE,
                total_errors INTEGER DEFAULT 0,
                errors_by_severity TEXT,
                errors_by_category TEXT,
                errors_by_module TEXT,
                resolved_errors INTEGER DEFAULT 0,
                unresolved_errors INTEGER DEFAULT 0,
                avg_resolution_time REAL,
                error_rate_per_hour REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 系统配置表
        conn.execute('''
            CREATE TABLE IF NOT EXISTS system_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                config_key TEXT NOT NULL UNIQUE,
                config_value TEXT,
                config_type TEXT,
                description TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        conn.commit()

    def _create_indexes(self, conn: sqlite3.Connection):
        """创建数据库索引"""
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_error_history_created_at ON error_history(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_error_history_severity ON error_history(severity)",
            "CREATE INDEX IF NOT EXISTS idx_error_history_category ON error_history(category)",
            "CREATE INDEX IF NOT EXISTS idx_error_history_module ON error_history(module)",
            "CREATE INDEX IF NOT EXISTS idx_error_history_resolved ON error_history(resolved)",
            "CREATE INDEX IF NOT EXISTS idx_error_statistics_daily_date ON error_statistics_daily(date)"
        ]

        for index_sql in indexes:
            try:
                conn.execute(index_sql)
            except sqlite3.Error as e:
                logger.warning(f"创建索引失败: {index_sql} - {e}")

        conn.commit()

    # ================ 核心CRUD操作 ================

    def save_error(self, error_record: ErrorRecord) -> bool:
        """
        保存错误记录到数据库

        Args:
            error_record: 错误记录对象

        Returns:
            bool: 保存是否成功
        """
        if not self.config.enabled:
            logger.debug("错误历史功能已禁用，跳过保存")
            return True

        try:
            with self._get_connection() as conn:
                # 准备数据
                data = {
                    'error_id': error_record.error_id,
                    'error_type': error_record.error_type,
                    'error_message': error_record.error_message,
                    'severity': error_record.severity.value,
                    'category': error_record.category.value,
                    'module': error_record.module,
                    'function': error_record.function,
                    'line_number': error_record.line_number,
                    'stack_trace': error_record.stack_trace,
                    'context': json.dumps(error_record.context, ensure_ascii=False) if error_record.context else None,
                    'user_context': json.dumps(error_record.user_context, ensure_ascii=False) if error_record.user_context else None,
                    'system_context': json.dumps(error_record.system_context, ensure_ascii=False) if error_record.system_context else None,
                    'resolved': error_record.resolved,
                    'resolved_at': error_record.resolved_at.isoformat() if error_record.resolved_at else None,
                    'resolution_method': error_record.resolution_method,
                    'resolution_time': error_record.resolution_time,
                    'retry_count': error_record.retry_count,
                    'max_retries': error_record.max_retries,
                    'tags': json.dumps(error_record.tags, ensure_ascii=False) if error_record.tags else None,
                    'metadata': json.dumps(error_record.metadata, ensure_ascii=False) if error_record.metadata else None
                }

                # 插入或更新
                conn.execute('''
                    INSERT OR REPLACE INTO error_history
                    (error_id, error_type, error_message, severity, category, module,
                     function, line_number, stack_trace, context, user_context, system_context,
                     resolved, resolved_at, resolution_method, resolution_time,
                     retry_count, max_retries, tags, metadata)
                    VALUES (:error_id, :error_type, :error_message, :severity, :category, :module,
                            :function, :line_number, :stack_trace, :context, :user_context, :system_context,
                            :resolved, :resolved_at, :resolution_method, :resolution_time,
                            :retry_count, :max_retries, :tags, :metadata)
                ''', data)

                # 立即提交事务
                conn.commit()

                # 更新日统计（在同一事务中）
                try:
                    self._update_daily_statistics(conn, error_record.created_at or datetime.now())
                    conn.commit()  # 确保日统计更新也提交
                except Exception as e:
                    logger.warning(f"更新日统计失败，但错误记录已保存: {e}")
                    conn.commit()  # 确保主事务提交

                logger.debug(f"错误记录已保存: {error_record.error_id}")
                return True

        except Exception as e:
            logger.error(f"保存错误记录失败: {e}")
            return False

    def get_error(self, error_id: str) -> Optional[ErrorRecord]:
        """
        根据错误ID获取错误记录

        Args:
            error_id: 错误ID

        Returns:
            ErrorRecord或None: 错误记录对象
        """
        try:
            with self._get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    "SELECT * FROM error_history WHERE error_id = ?",
                    (error_id,)
                )
                row = cursor.fetchone()

                if row:
                    return self._row_to_error_record(row)

        except Exception as e:
            logger.error(f"获取错误记录失败: {e}")

        return None

    def update_error(self, error_record: ErrorRecord) -> bool:
        """
        更新错误记录

        Args:
            error_record: 错误记录对象

        Returns:
            bool: 更新是否成功
        """
        if not error_record.id and not error_record.error_id:
            logger.error("更新错误记录需要ID或error_id")
            return False

        try:
            with self._get_connection() as conn:
                # 准备更新数据
                data = {
                    'error_type': error_record.error_type,
                    'error_message': error_record.error_message,
                    'severity': error_record.severity.value,
                    'category': error_record.category.value,
                    'module': error_record.module,
                    'function': error_record.function,
                    'line_number': error_record.line_number,
                    'stack_trace': error_record.stack_trace,
                    'context': json.dumps(error_record.context, ensure_ascii=False) if error_record.context else None,
                    'user_context': json.dumps(error_record.user_context, ensure_ascii=False) if error_record.user_context else None,
                    'system_context': json.dumps(error_record.system_context, ensure_ascii=False) if error_record.system_context else None,
                    'resolved': error_record.resolved,
                    'resolved_at': error_record.resolved_at.isoformat() if error_record.resolved_at else None,
                    'resolution_method': error_record.resolution_method,
                    'resolution_time': error_record.resolution_time,
                    'retry_count': error_record.retry_count,
                    'max_retries': error_record.max_retries,
                    'tags': json.dumps(error_record.tags, ensure_ascii=False) if error_record.tags else None,
                    'metadata': json.dumps(error_record.metadata, ensure_ascii=False) if error_record.metadata else None
                }

                # 构建更新条件
                if error_record.id:
                    data['id'] = error_record.id
                    where_clause = "id = :id"
                else:
                    data['error_id'] = error_record.error_id
                    where_clause = "error_id = :error_id"

                # 执行更新
                cursor = conn.execute(f'''
                    UPDATE error_history SET
                        error_type = :error_type,
                        error_message = :error_message,
                        severity = :severity,
                        category = :category,
                        module = :module,
                        function = :function,
                        line_number = :line_number,
                        stack_trace = :stack_trace,
                        context = :context,
                        user_context = :user_context,
                        system_context = :system_context,
                        resolved = :resolved,
                        resolved_at = :resolved_at,
                        resolution_method = :resolution_method,
                        resolution_time = :resolution_time,
                        retry_count = :retry_count,
                        max_retries = :max_retries,
                        tags = :tags,
                        metadata = :metadata
                    WHERE {where_clause}
                ''', data)

                conn.commit()

                if cursor.rowcount > 0:
                    logger.debug(f"错误记录已更新: {error_record.error_id or error_record.id}")
                    return True
                else:
                    logger.warning(f"未找到要更新的错误记录: {error_record.error_id or error_record.id}")
                    return False

        except Exception as e:
            logger.error(f"更新错误记录失败: {e}")
            return False

    def delete_error(self, error_id: str) -> bool:
        """
        删除错误记录

        Args:
            error_id: 错误ID

        Returns:
            bool: 删除是否成功
        """
        try:
            with self._get_connection() as conn:
                cursor = conn.execute(
                    "DELETE FROM error_history WHERE error_id = ?",
                    (error_id,)
                )
                conn.commit()

                if cursor.rowcount > 0:
                    logger.debug(f"错误记录已删除: {error_id}")
                    return True
                else:
                    logger.warning(f"未找到要删除的错误记录: {error_id}")
                    return False

        except Exception as e:
            logger.error(f"删除错误记录失败: {e}")
            return False

    # ================ 查询方法 ================

    def query_errors(self, filters: Dict[str, Any] = None, limit: int = 100,
                    offset: int = 0, order_by: str = "created_at DESC") -> List[ErrorRecord]:
        """
        根据条件查询错误记录

        Args:
            filters: 查询过滤条件
            limit: 返回记录数量限制
            offset: 偏移量
            order_by: 排序方式

        Returns:
            List[ErrorRecord]: 错误记录列表
        """
        filters = filters or {}

        try:
            with self._get_connection() as conn:
                conn.row_factory = sqlite3.Row

                # 构建查询条件
                where_clauses = []
                query_params = []  # 使用位置参数列表

                # 严重程度过滤
                if 'severity' in filters:
                    severity = filters['severity']
                    if isinstance(severity, ErrorSeverity):
                        where_clauses.append("severity = ?")
                        query_params.append(severity.value)
                    elif isinstance(severity, list):
                        placeholders = ','.join('?' * len(severity))
                        where_clauses.append(f"severity IN ({placeholders})")
                        for s in severity:
                            query_params.append(s.value if hasattr(s, 'value') else str(s))

                # 分类过滤
                if 'category' in filters:
                    category = filters['category']
                    if isinstance(category, ErrorCategory):
                        where_clauses.append("category = ?")
                        query_params.append(category.value)
                    elif isinstance(category, list):
                        placeholders = ','.join('?' * len(category))
                        where_clauses.append(f"category IN ({placeholders})")
                        for c in category:
                            query_params.append(c.value if hasattr(c, 'value') else str(c))

                # 模块过滤
                if 'module' in filters:
                    where_clauses.append("module = ?")
                    query_params.append(filters['module'])

                # 解决状态过滤
                if 'resolved' in filters:
                    where_clauses.append("resolved = ?")
                    query_params.append(filters['resolved'])

                # 时间范围过滤
                if 'start_date' in filters:
                    where_clauses.append("date(created_at) >= ?")
                    query_params.append(filters['start_date'].isoformat())

                if 'end_date' in filters:
                    where_clauses.append("date(created_at) <= ?")
                    query_params.append(filters['end_date'].isoformat())

                # 错误类型过滤
                if 'error_type' in filters:
                    where_clauses.append("error_type LIKE ?")
                    query_params.append(f"%{filters['error_type']}%")

                # 错误消息过滤
                if 'error_message' in filters:
                    where_clauses.append("error_message LIKE ?")
                    query_params.append(f"%{filters['error_message']}%")

                # 构建查询语句
                where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"
                query = f"""
                    SELECT * FROM error_history
                    WHERE {where_clause}
                    ORDER BY {order_by}
                    LIMIT ? OFFSET ?
                """
                query_params.extend([limit, offset])

                cursor = conn.execute(query, query_params)
                rows = cursor.fetchall()

                # 转换为ErrorRecord对象
                errors = []
                for row in rows:
                    errors.append(self._row_to_error_record(row))

                return errors

        except Exception as e:
            logger.error(f"查询错误记录失败: {e}")
            return []

    def get_recent_errors(self, limit: int = 50) -> List[ErrorRecord]:
        """
        获取最近的错误记录

        Args:
            limit: 返回记录数量

        Returns:
            List[ErrorRecord]: 最近的错误记录列表
        """
        return self.query_errors(limit=limit)

    def get_errors_by_date_range(self, start_date: date, end_date: date) -> List[ErrorRecord]:
        """
        获取指定日期范围内的错误记录

        Args:
            start_date: 开始日期
            end_date: 结束日期

        Returns:
            List[ErrorRecord]: 错误记录列表
        """
        return self.query_errors(
            filters={'start_date': start_date, 'end_date': end_date},
            limit=10000  # 较大的限制
        )

    # ================ 统计方法 ================

    def get_statistics(self, date_range: Tuple[date, date] = None) -> Dict[str, Any]:
        """
        获取错误统计信息

        Args:
            date_range: 日期范围 (开始日期, 结束日期)，如果为None则统计全部

        Returns:
            Dict[str, Any]: 统计信息字典
        """
        try:
            with self._get_connection() as conn:
                # 设置日期范围条件
                date_condition = ""
                date_params = []

                if date_range:
                    start_date, end_date = date_range
                    date_condition = "AND date(created_at) BETWEEN ? AND ?"
                    date_params = [start_date.isoformat(), end_date.isoformat()]

                # 基本统计
                cursor = conn.execute(f'''
                    SELECT
                        COUNT(*) as total_errors,
                        SUM(CASE WHEN resolved = 1 THEN 1 ELSE 0 END) as resolved_errors,
                        SUM(CASE WHEN resolved = 0 THEN 1 ELSE 0 END) as unresolved_errors,
                        AVG(CASE WHEN resolved = 1 THEN resolution_time ELSE NULL END) as avg_resolution_time
                    FROM error_history
                    WHERE 1=1 {date_condition}
                ''', date_params)

                basic_stats = cursor.fetchone()

                # 按严重程度统计
                cursor = conn.execute(f'''
                    SELECT severity, COUNT(*) as count
                    FROM error_history
                    WHERE 1=1 {date_condition}
                    GROUP BY severity
                    ORDER BY count DESC
                ''', date_params)

                severity_stats = {row[0]: row[1] for row in cursor.fetchall()}

                # 按分类统计
                cursor = conn.execute(f'''
                    SELECT category, COUNT(*) as count
                    FROM error_history
                    WHERE 1=1 {date_condition}
                    GROUP BY category
                    ORDER BY count DESC
                ''', date_params)

                category_stats = {row[0]: row[1] for row in cursor.fetchall()}

                # 按模块统计
                cursor = conn.execute(f'''
                    SELECT module, COUNT(*) as count
                    FROM error_history
                    WHERE module IS NOT NULL AND module != ''
                    AND 1=1 {date_condition}
                    GROUP BY module
                    ORDER BY count DESC
                    LIMIT 20
                ''', date_params)

                module_stats = {row[0]: row[1] for row in cursor.fetchall()}

                # 计算错误率
                if date_range:
                    days = (end_date - start_date).days + 1
                    error_rate_per_day = basic_stats[0] / days if days > 0 else 0
                    error_rate_per_hour = error_rate_per_day / 24
                else:
                    # 计算最近30天的错误率
                    cursor = conn.execute('''
                        SELECT COUNT(*) as count
                        FROM error_history
                        WHERE created_at >= datetime('now', '-30 days')
                    ''')
                    recent_count = cursor.fetchone()[0]
                    error_rate_per_hour = recent_count / (30 * 24)

                return {
                    'total_errors': basic_stats[0] or 0,
                    'resolved_errors': basic_stats[1] or 0,
                    'unresolved_errors': basic_stats[2] or 0,
                    'avg_resolution_time': basic_stats[3],
                    'error_rate_per_hour': error_rate_per_hour,
                    'errors_by_severity': severity_stats,
                    'errors_by_category': category_stats,
                    'errors_by_module': module_stats,
                    'date_range': f"{start_date.isoformat()} to {end_date.isoformat()}" if date_range else None
                }

        except Exception as e:
            logger.error(f"获取统计信息失败: {e}")
            return {}

    def get_daily_statistics(self, target_date: date = None) -> Optional[DailyStatistics]:
        """
        获取指定日期的日统计信息

        Args:
            target_date: 目标日期，如果为None则使用今天

        Returns:
            DailyStatistics或None: 日统计信息
        """
        if target_date is None:
            target_date = date.today()

        try:
            with self._get_connection() as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    "SELECT * FROM error_statistics_daily WHERE date = ?",
                    (target_date.isoformat(),)
                )
                row = cursor.fetchone()

                if row:
                    return DailyStatistics(
                        date=target_date,
                        total_errors=row['total_errors'],
                        errors_by_severity=json.loads(row['errors_by_severity'] or '{}'),
                        errors_by_category=json.loads(row['errors_by_category'] or '{}'),
                        errors_by_module=json.loads(row['errors_by_module'] or '{}'),
                        resolved_errors=row['resolved_errors'],
                        unresolved_errors=row['unresolved_errors'],
                        avg_resolution_time=row['avg_resolution_time'],
                        error_rate_per_hour=row['error_rate_per_hour']
                    )
                else:
                    # 如果没有现成统计，实时计算
                    return self._calculate_daily_statistics(target_date)

        except Exception as e:
            logger.error(f"获取日统计信息失败: {e}")

        return None

    def _calculate_daily_statistics(self, target_date: date) -> DailyStatistics:
        """计算指定日期的日统计信息"""
        try:
            with self._get_connection() as conn:
                # 设置row_factory为字典模式
                conn.row_factory = sqlite3.Row

                # 获取当日错误
                cursor = conn.execute('''
                    SELECT * FROM error_history
                    WHERE date(created_at) = ?
                ''', (target_date.isoformat(),))

                errors = []
                for row in cursor.fetchall():
                    errors.append(self._row_to_error_record(row))

                if not errors:
                    return DailyStatistics(date=target_date)

                # 计算统计信息
                total_errors = len(errors)
                resolved_errors = sum(1 for e in errors if e.resolved)
                unresolved_errors = total_errors - resolved_errors

                # 按严重程度统计
                severity_counts = {}
                for error in errors:
                    severity = error.severity.value
                    severity_counts[severity] = severity_counts.get(severity, 0) + 1

                # 按分类统计
                category_counts = {}
                for error in errors:
                    category = error.category.value
                    category_counts[category] = category_counts.get(category, 0) + 1

                # 按模块统计
                module_counts = {}
                for error in errors:
                    if error.module:
                        module_counts[error.module] = module_counts.get(error.module, 0) + 1

                # 平均解决时间
                resolved_times = [e.resolution_time for e in errors if e.resolved and e.resolution_time]
                avg_resolution_time = sum(resolved_times) / len(resolved_times) if resolved_times else None

                # 错误率（每小时）
                error_rate_per_hour = total_errors / 24

                stats = DailyStatistics(
                    date=target_date,
                    total_errors=total_errors,
                    errors_by_severity=severity_counts,
                    errors_by_category=category_counts,
                    errors_by_module=module_counts,
                    resolved_errors=resolved_errors,
                    unresolved_errors=unresolved_errors,
                    avg_resolution_time=avg_resolution_time,
                    error_rate_per_hour=error_rate_per_hour
                )

                # 保存到数据库
                self._save_daily_statistics(conn, stats)

                # 提交事务
                conn.commit()

                return stats

        except Exception as e:
            logger.error(f"计算日统计信息失败: {e}")
            return DailyStatistics(date=target_date)

    def _update_daily_statistics(self, conn: sqlite3.Connection, error_datetime: datetime):
        """更新日统计信息"""
        try:
            error_date = error_datetime.date()

            # 检查是否已有统计记录
            cursor = conn.execute(
                "SELECT id FROM error_statistics_daily WHERE date = ?",
                (error_date.isoformat(),)
            )

            if cursor.fetchone():
                # 更新现有记录
                stats = self._calculate_daily_statistics(error_date)
                self._save_daily_statistics(conn, stats)
            else:
                # 创建新记录
                stats = self._calculate_daily_statistics(error_date)
                self._save_daily_statistics(conn, stats)

        except Exception as e:
            logger.warning(f"更新日统计信息失败: {e}")

    def _save_daily_statistics(self, conn: sqlite3.Connection, stats: DailyStatistics):
        """保存日统计信息到数据库"""
        try:
            data = {
                'date': stats.date.isoformat(),
                'total_errors': stats.total_errors,
                'errors_by_severity': json.dumps(stats.errors_by_severity, ensure_ascii=False),
                'errors_by_category': json.dumps(stats.errors_by_category, ensure_ascii=False),
                'errors_by_module': json.dumps(stats.errors_by_module, ensure_ascii=False),
                'resolved_errors': stats.resolved_errors,
                'unresolved_errors': stats.unresolved_errors,
                'avg_resolution_time': stats.avg_resolution_time,
                'error_rate_per_hour': stats.error_rate_per_hour
            }

            conn.execute('''
                INSERT OR REPLACE INTO error_statistics_daily
                (date, total_errors, errors_by_severity, errors_by_category,
                 errors_by_module, resolved_errors, unresolved_errors,
                 avg_resolution_time, error_rate_per_hour)
                VALUES (:date, :total_errors, :errors_by_severity, :errors_by_category,
                        :errors_by_module, :resolved_errors, :unresolved_errors,
                        :avg_resolution_time, :error_rate_per_hour)
            ''', data)

            # 提交事务
            conn.commit()

        except Exception as e:
            logger.warning(f"保存日统计信息失败: {e}")

    # ================ 管理方法 ================

    def cleanup_old_errors(self, days: int = None) -> int:
        """
        清理指定天数前的错误记录

        Args:
            days: 保留天数，如果为None则使用配置值

        Returns:
            int: 删除的记录数量
        """
        if days is None:
            days = self.config.retention_days

        try:
            with self._get_connection() as conn:
                # 计算截止日期
                cutoff_date = (datetime.now() - timedelta(days=days)).isoformat()

                # 删除旧记录
                cursor = conn.execute(
                    "DELETE FROM error_history WHERE created_at < ?",
                    (cutoff_date,)
                )

                deleted_count = cursor.rowcount
                conn.commit()

                # 清理相关的日统计（可选）
                if deleted_count > 0:
                    self._cleanup_old_statistics(conn, days)

                logger.info(f"已清理 {deleted_count} 条过期错误记录（{days}天前）")
                return deleted_count

        except Exception as e:
            logger.error(f"清理过期错误记录失败: {e}")
            return 0

    def _cleanup_old_statistics(self, conn: sqlite3.Connection, days: int):
        """清理过期的日统计记录"""
        try:
            cutoff_date = (date.today() - timedelta(days=days)).isoformat()
            cursor = conn.execute(
                "DELETE FROM error_statistics_daily WHERE date < ?",
                (cutoff_date,)
            )
            deleted_stats = cursor.rowcount
            if deleted_stats > 0:
                logger.info(f"已清理 {deleted_stats} 条过期统计记录")
        except Exception as e:
            logger.warning(f"清理过期统计记录失败: {e}")

    def optimize_database(self) -> bool:
        """
        优化数据库性能

        Returns:
            bool: 优化是否成功
        """
        try:
            with self._get_connection() as conn:
                # 执行VACUUM
                conn.execute("VACUUM")

                # 重新分析表统计信息
                conn.execute("ANALYZE")

                # 重新索引
                conn.execute("REINDEX")

                conn.commit()

                logger.info("数据库优化完成")
                return True

        except Exception as e:
            logger.error(f"数据库优化失败: {e}")
            return False

    def backup_database(self, backup_path: str = None) -> bool:
        """
        备份数据库

        Args:
            backup_path: 备份文件路径，如果为None则使用默认路径

        Returns:
            bool: 备份是否成功
        """
        if backup_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"data/error_history_backup_{timestamp}.db"

        try:
            # 使用SQLite的VACUUM INTO命令（SQLite 3.27+）
            with self._get_connection() as conn:
                conn.execute(f"VACUUM INTO ?", (backup_path,))

            logger.info(f"数据库备份完成: {backup_path}")
            return True

        except Exception as e:
            logger.error(f"数据库备份失败: {e}")
            return False

    def export_data(self, file_path: str, format: str = None,
                   filters: Dict[str, Any] = None) -> bool:
        """
        导出错误数据

        Args:
            file_path: 导出文件路径
            format: 导出格式 (json, csv, xlsx)，如果为None则使用配置默认值
            filters: 导出过滤条件

        Returns:
            bool: 导出是否成功
        """
        if format is None:
            format = self.config.default_format

        if format not in self.config.supported_formats:
            logger.error(f"不支持的导出格式: {format}")
            return False

        try:
            # 获取要导出的数据
            errors = self.query_errors(
                filters=filters,
                limit=self.config.max_export_rows
            )

            if not errors:
                logger.warning("没有找到要导出的数据")
                return False

            # 根据格式导出
            if format == 'json':
                self._export_json(errors, file_path)
            elif format == 'csv':
                self._export_csv(errors, file_path)
            elif format == 'xlsx':
                self._export_xlsx(errors, file_path)

            logger.info(f"数据导出完成: {file_path} (格式: {format}, 记录数: {len(errors)})")
            return True

        except Exception as e:
            logger.error(f"数据导出失败: {e}")
            return False

    def _export_json(self, errors: List[ErrorRecord], file_path: str):
        """导出为JSON格式"""
        data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'total_records': len(errors),
                'format': 'json'
            },
            'errors': [error.to_dict() for error in errors]
        }

        import json
        with builtins.open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def _export_csv(self, errors: List[ErrorRecord], file_path: str):
        """导出为CSV格式"""
        import csv

        with builtins.open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # 写入表头
            headers = [
                'error_id', 'error_type', 'error_message', 'severity', 'category',
                'module', 'function', 'line_number', 'created_at', 'resolved',
                'resolution_method', 'resolution_time', 'retry_count'
            ]
            writer.writerow(headers)

            # 写入数据
            for error in errors:
                row = [
                    error.error_id,
                    error.error_type,
                    error.error_message,
                    error.severity.value,
                    error.category.value,
                    error.module,
                    error.function,
                    error.line_number,
                    error.created_at.isoformat() if error.created_at else '',
                    error.resolved,
                    error.resolution_method,
                    error.resolution_time,
                    error.retry_count
                ]
                writer.writerow(row)

    def _export_xlsx(self, errors: List[ErrorRecord], file_path: str):
        """导出为Excel格式"""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "错误历史"

            # 写入表头
            headers = [
                '错误ID', '错误类型', '错误消息', '严重程度', '分类',
                '模块', '函数', '行号', '创建时间', '已解决',
                '解决方法', '解决时间', '重试次数'
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row, error in enumerate(errors, 2):
                ws.cell(row=row, column=1, value=error.error_id)
                ws.cell(row=row, column=2, value=error.error_type)
                ws.cell(row=row, column=3, value=error.error_message)
                ws.cell(row=row, column=4, value=error.severity.value)
                ws.cell(row=row, column=5, value=error.category.value)
                ws.cell(row=row, column=6, value=error.module)
                ws.cell(row=row, column=7, value=error.function)
                ws.cell(row=row, column=8, value=error.line_number)
                ws.cell(row=row, column=9, value=error.created_at.isoformat() if error.created_at else '')
                ws.cell(row=row, column=10, value=error.resolved)
                ws.cell(row=row, column=11, value=error.resolution_method)
                ws.cell(row=row, column=12, value=error.resolution_time)
                ws.cell(row=row, column=13, value=error.retry_count)

            wb.save(file_path)

        except ImportError:
            logger.error("导出Excel格式需要安装openpyxl库")
            raise

    # ================ 工具方法 ================

    def _row_to_error_record(self, row) -> ErrorRecord:
        """将数据库行转换为ErrorRecord对象"""
        return ErrorRecord(
            id=row['id'],
            error_id=row['error_id'],
            error_type=row['error_type'],
            error_message=row['error_message'],
            severity=ErrorSeverity(row['severity']),
            category=ErrorCategory(row['category']),
            module=row['module'],
            function=row['function'],
            line_number=row['line_number'],
            stack_trace=row['stack_trace'],
            context=json.loads(row['context']) if row['context'] else None,
            user_context=json.loads(row['user_context']) if row['user_context'] else None,
            system_context=json.loads(row['system_context']) if row['system_context'] else None,
            created_at=datetime.fromisoformat(row['created_at']),
            resolved=bool(row['resolved']),
            resolved_at=datetime.fromisoformat(row['resolved_at']) if row['resolved_at'] else None,
            resolution_method=row['resolution_method'],
            resolution_time=row['resolution_time'],
            retry_count=row['retry_count'],
            max_retries=row['max_retries'],
            tags=json.loads(row['tags']) if row['tags'] else None,
            metadata=json.loads(row['metadata']) if row['metadata'] else None
        )

    def get_database_info(self) -> Dict[str, Any]:
        """
        获取数据库信息

        Returns:
            Dict[str, Any]: 数据库信息
        """
        try:
            with self._get_connection() as conn:
                # 获取表信息
                cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [row[0] for row in cursor.fetchall()]

                # 获取记录数量
                table_counts = {}
                for table in tables:
                    cursor = conn.execute(f"SELECT COUNT(*) FROM {table}")
                    count = cursor.fetchone()[0]
                    table_counts[table] = count

                # 获取数据库文件大小
                db_size = self.db_path.stat().st_size if self.db_path.exists() else 0

                return {
                    'database_path': str(self.db_path),
                    'database_size': db_size,
                    'tables': tables,
                    'table_counts': table_counts,
                    'connection_pool_size': len(self._connection_pool)
                }

        except Exception as e:
            logger.error(f"获取数据库信息失败: {e}")
            return {}

    def shutdown(self):
        """关闭管理器，清理资源"""
        try:
            # 停止调度与监听
            try:
                self._stop_cleanup_scheduler()
            except Exception:
                pass
            try:
                self._stop_config_watcher()
            except Exception:
                pass
            try:
                self._try_watchdog_stop()
            except Exception:
                pass
            # 关闭所有连接
            with self._pool_lock:
                for conn in self._connection_pool.values():
                    try:
                        conn.close()
                    except Exception as e:
                        logger.warning(f"关闭数据库连接失败: {e}")

                self._connection_pool.clear()

            logger.info("错误历史管理器已关闭")

        except Exception as e:
            logger.error(f"关闭错误历史管理器失败: {e}")
